"""
Процессор экспорта данных.

Формирует результаты в виде списка или Excel-файла.
"""

import os
import tempfile
from datetime import datetime
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font
from config import Config
from models.user import TelegramUser

class ExportProcessor:
    """Процессор для формирования результатов экспорта."""

    def __init__(self):
        pass

    def generate_export(self, users: List[TelegramUser]) -> str:
        """Генерация экспорта в зависимости от количества пользователей."""
        if len(users) < Config.EXPORT_THRESHOLD:
            return self._generate_text_export(users)
        else:
            return self._generate_excel_export(users)

    def _generate_text_export(self, users: List[TelegramUser]) -> str:
        """Генерация текстового экспорта для небольшого количества пользователей."""
        try:
            # Создание временного файла
            temp_file = tempfile.NamedTemporaryFile(
                mode='w',
                suffix='.txt',
                encoding='utf-8',
                delete=False
            )

            # Запись заголовка
            temp_file.write(f"Список участников чата ({len(users)} пользователей)\n\n")

            # Запись пользователей
            for i, user in enumerate(users, 1):
                temp_file.write(f"{i}. {user.display_name}\n")
                if user.user_id:
                    temp_file.write(f"   Telegram ID: {user.user_id}\n")
                if user.phone_number:
                    temp_file.write(f"   Телефон: {user.phone_number}\n")

            temp_file.close()
            return temp_file.name

        except Exception as e:
            raise Exception(f"Ошибка при генерации текстового экспорта: {str(e)}")

    def _generate_excel_export(self, users: List[TelegramUser]) -> str:
        """Генерация Excel экспорта для большого количества пользователей."""
        try:
            # Дедупликация пользователей по telegram_id (user_id)
            unique_users_dict = {}
            for user in users:
                if user.user_id not in unique_users_dict:
                    unique_users_dict[user.user_id] = user
            
            users = list(unique_users_dict.values())
            
            # Создание временного файла
            temp_file = tempfile.NamedTemporaryFile(
                suffix='.xlsx',
                delete=False
            )
            temp_file.close()

            # Создание книги Excel
            wb = Workbook()
            
            # Удаляем дефолтный лист
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            # Разделение пользователей на категории
            regular_users = []  # Обычные участники (включая упомянутых)
            channel_users = []  # Пользователи с каналами
            
            for user in users:
                if user.has_channel is True:  # Явно True, не None
                    channel_users.append(user)
                else:
                    # Все остальные пользователи (включая упомянутых) идут в общий список
                    regular_users.append(user)

            # Создание вкладки "Участники" (включая упомянутых)
            if regular_users:
                ws_participants = wb.create_sheet("Участники")
                self._write_users_to_sheet(ws_participants, regular_users)

            # Создание вкладки "Каналы"
            if channel_users:
                ws_channels = wb.create_sheet("Каналы")
                self._write_users_to_sheet(ws_channels, channel_users)

            # Если нет данных, создаем хотя бы одну вкладку
            if not regular_users and not channel_users:
                ws = wb.create_sheet("Участники")
                self._write_users_to_sheet(ws, users)

            # Сохранение файла
            wb.save(temp_file.name)
            return temp_file.name

        except Exception as e:
            raise Exception(f"Ошибка при генерации Excel экспорта: {str(e)}")

    def _write_users_to_sheet(self, ws, users: List[TelegramUser]):
        """Запись пользователей на лист Excel."""
        # Запись заголовков
        headers = Config.EXCEL_COLUMNS
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)

        # Запись данных
        # Структура: Telegram ID, Имя, Телефон, Упоминание, Дата 1-го сообщения, ID 1-го сообщения, Дата 1-й реакции, Эмодзи 1-й реакции, Бот
        for row_num, user in enumerate(users, 2):
            try:
                # Telegram ID
                ws.cell(row=row_num, column=1, value=user.user_id)
                
                # Имя (объединенное имя и фамилия)
                full_name = user.full_name  # Используем свойство full_name из модели
                ws.cell(row=row_num, column=2, value=full_name)
                
                # Телефон
                ws.cell(row=row_num, column=3, value=user.phone_number)
                
                # Упоминание
                ws.cell(row=row_num, column=4, value=user.mention)
                
                # Дата первого сообщения
                first_msg_date = None
                if user.first_message_date:
                    if isinstance(user.first_message_date, datetime):
                        first_msg_date = user.first_message_date.strftime('%Y-%m-%d %H:%M:%S')
                    else:
                        first_msg_date = str(user.first_message_date)
                ws.cell(row=row_num, column=5, value=first_msg_date)
                
                # ID первого сообщения
                ws.cell(row=row_num, column=6, value=user.first_message_id)
                
                # Дата первой реакции
                first_reaction_date = None
                if user.first_reaction_date:
                    if isinstance(user.first_reaction_date, datetime):
                        first_reaction_date = user.first_reaction_date.strftime('%Y-%m-%d %H:%M:%S')
                    else:
                        first_reaction_date = str(user.first_reaction_date)
                ws.cell(row=row_num, column=7, value=first_reaction_date)
                
                # Эмодзи первой реакции
                ws.cell(row=row_num, column=8, value=user.first_reaction_emoji)
                
                # Бот
                ws.cell(row=row_num, column=9, value="Да" if user.is_bot else "Нет")
                
            except Exception as e:
                # Если возникает ошибка для конкретного пользователя, записываем хотя бы ID
                ws.cell(row=row_num, column=1, value=user.user_id if hasattr(user, 'user_id') else "Ошибка")
                ws.cell(row=row_num, column=2, value=f"Ошибка: {str(e)}")
                ws.cell(row=row_num, column=3, value=None)
                ws.cell(row=row_num, column=4, value=None)
                ws.cell(row=row_num, column=5, value=None)
                ws.cell(row=row_num, column=6, value=None)
                ws.cell(row=row_num, column=7, value=None)
                ws.cell(row=row_num, column=8, value=None)
                ws.cell(row=row_num, column=9, value="Нет")

        # Авторазмер колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width if adjusted_width > 10 else 15

    def cleanup(self, file_path: str):
        """Удаление временного файла."""
        if os.path.exists(file_path):
            os.remove(file_path)